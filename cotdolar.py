#!/usr/bin/env python3
#
# dolarexchangerate.py by Regis Froemming
# created 2022-02-07
# updated 2022-03-06

#from tkinter.ttk import Style
import requests 
from tkinter import StringVar, Label, Button, Tk, messagebox
import tkinter.font as tkFont 
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
from tkcalendar import Calendar
import sys
import smtplib, ssl


import os.path

class InitFile:
        # file location 
        dest_filename = 'data.xlsx'
        wb = Workbook()

        # check if file exist, if not, create new file 
        if not os.path.isfile(dest_filename):
            ws = wb.active
            ws.title = "Data"
            ws.append({'A':'Date','B':'Time','C':'Value','D':'API'})
            wb.save(filename = dest_filename)

        # load excel sheet
        wb = load_workbook(filename = dest_filename)    
        
        # active sheet
        ws = wb['Data']

        # read last data to an array
        last_date = False
        exchangeoftheday = []
        for col in ws['A']:
            if col.value!='Date':
                d1 = datetime.strptime(col.value, '%d/%m/%Y')
                if not last_date:
                    last_date=d1
                if d1!=last_date:
                    exchangeoftheday.append([last_date.strftime('%d/%m/%Y'),last_value])
                last_date = d1 
                last_value = col.offset(column=2).value
            if (len(exchangeoftheday))>4:
                del exchangeoftheday[0]      

class MainWindow:
    def __init__(self, master):
        
        # configure the root window
        self.master = master
        master.title("Canadian dolar to Real exchange")
        master.geometry("400x400")
        
        # labels
        labelFontStyle = tkFont.Font(family="Lucida Grande", size=15)
        self.lbl_lastupdatedate_text = StringVar()
        self.lbl_lastupdatedate_text.set("Date: ")
        self.lbl_lastupdatedate = Label(master, textvariable=self.lbl_lastupdatedate_text, font=labelFontStyle)
        self.lbl_lastupdatedate.place(x=30, y=20)

        self.lbl_lastupdatetime_text = StringVar()
        self.lbl_lastupdatetime_text.set("Time: ")
        self.lbl_lastupdatetime = Label(master, textvariable=self.lbl_lastupdatetime_text, font=labelFontStyle)
        self.lbl_lastupdatetime.place(x=30, y=60)

        self.lbl_lastupdatevalue_text = StringVar()
        self.lbl_lastupdatevalue_text.set("Last Exchange Rate: ")
        self.lbl_lastupdatevalue = Label(master, textvariable=self.lbl_lastupdatevalue_text, font=labelFontStyle)
        self.lbl_lastupdatevalue.place(x=30, y=100)

        # buttons
        buttonFontStyle = tkFont.Font(family="Lucida Grande", weight="bold", size=11)
        self.update = Button(master, text = 'Update',command = self.Update,font=buttonFontStyle)
        self.update.place(x= 30, y=160)
        self.graph = Button(master, text = 'Graph',command = self.CalSel,font=buttonFontStyle)
        self.graph.place(x= 120, y=160)
        self.read = Button(master, text = 'Read',command = self.ReadLastValue, font=buttonFontStyle)
        self.read.place(x= 210, y=160)
        

        # self.now = datetime.now()
        # self.cal = Calendar(self.master, selectmode = 'day',
        #        year = self.now.year, month = self.now.month,
        #        day = self.now.day,maxdate=self.now)

        # self.cal.bind('<<CalendarSelected>>',print('xxx'))
        
        #self.ReadLastValue()
    
    def Update(self):
        # call function to update data
        rtr=UpdateCmd()
        if rtr[0]!='error':
            # update labels
            self.lbl_lastupdatedate_text.set("Date: "+rtr[0])
            self.lbl_lastupdatetime_text.set("Time: "+rtr[1])
            self.lbl_lastupdatevalue_text.set("Current Exchange Rate: "+str(rtr[2]))
        else:
            messagebox.showerror("Error", rtr[1])
        
    def ReadLastValue(self):
        # check last line from the sheet and read line 
        self.lastRow=len(InitFile.ws['A'])
        
        self.lbl_lastupdatedate_text.set("Date: "+InitFile.ws['A'+str(self.lastRow)].value)
        self.lbl_lastupdatetime_text.set("Time: "+InitFile.ws['B'+str(self.lastRow)].value)
        self.lbl_lastupdatevalue_text.set("Last Exchange Rate: "+str(InitFile.ws['C'+str(self.lastRow)].value))
        # last_date = False
        # exchangeoftheday = []
        # for col in InitFile.ws['A']:
        #     if col.value!='Date':
        #         d1 = datetime.strptime(col.value, '%d/%m/%Y')
        #         if not last_date:
        #             last_date=d1
        #         if d1!=last_date:
        #             exchangeoftheday.append([last_date.strftime('%d/%m/%Y'),last_value])
        #         last_date = d1 
        #         last_value = col.offset(column=2).value
        #     if (len(exchangeoftheday))>4:
        #         del exchangeoftheday[0]    
        
        posy = 220
        self.lbl1 = Label(self.master, text='Last values:',font='Arial')
        self.lbl1.place(x=30, y=posy)
        for r in reversed(InitFile.exchangeoftheday):
            posy += 30
            self.lbl1 = Label(self.master, text='Date: '+r[0]+'  Value: '+str(r[1]))
            self.lbl1.place(x=30, y=posy)

    def CalSel(self):
        self.de=Calendar(self.master,maxdate=datetime.now(),date_pattern='dd/mm/yyyy')  
        self.de.place(x= 120, y=195)  
        self.de.bind("<<CalendarSelected>>", self.Graph)  
    
    def Graph(self, event):
        w = event.widget
        date = w.get_date()
        #print('Selected Date:{}'.format(date))
        
        # closes calendar
        self.de.destroy()
        x = []
        y = []
        for col in InitFile.ws['A']:
            if col.value!='Date':
                #now = datetime.now()
                d1 = datetime.strptime(col.value, '%d/%m/%Y')
                d2 = datetime.strptime(date, '%d/%m/%Y')
                if d1==d2:
                    x.append(col.offset(column=1).value)
                    y.append(col.offset(column=2).value)
        # figure = plt.figure(figsize=(6,4))
        plt.style.use('seaborn')
        plt.rcParams['toolbar'] = 'None'
        #plt.title("Canadian Dollar to Real")
        fig = plt.figure("Dollar Exchange rates")
        plt.xticks(rotation=45, fontsize=9)
        plt.plot(x, y)
        plt.show()

def UpdateCmd():
    try:
        # stores current date and time in variables
        now = datetime.now()
        current_date = now.strftime("%d/%m/%Y")
        current_time = now.strftime("%H:%M:%S")
        
        # API to request exchange rate (sometime this API is offline)
        req = requests.get('https://free.currconv.com/api/v7/convert?q=CAD_BRL&apiKey=afbab7b7c9b75b72a2f0')
        if req:
            rate = req.json()
            last_value = rate['results']['CAD_BRL']['val']
            api = 'currentconverse.com'
        else:
            req = requests.get('https://economia.awesomeapi.com.br/last/CAD-BRL')
            rate = req.json()
            last_value = float(rate['CADBRL']['ask'])
            api = 'awesomeapi.ca'
        return WriteData(current_date,current_time,last_value,api)

    except:
        msg="Error trying to connect to API server."
        print(msg)
        return 'error',msg

def CheckChange():
    # value of percentage to check
    perc_check = 2 
    lastRow=len(InitFile.ws['C'])
    last_value=InitFile.ws['C'+str(lastRow)].value
    previus_value=InitFile.ws['C'+str(lastRow-1)].value
    # Compare current value to last value read
    if previus_value!='Value':
        percent = last_value/previus_value
        if percent >= (1+perc_check/100):
            msg = 'Alert - dollar going up - compared to last value +{:.2f}%'.format((percent-1)*100)
            msg = msg+'\nCurrent value: R$ '+str(last_value)
            sub='Dollar Exchange notification'
            Sendmail(sub,msg)
        elif percent <= (1-perc_check/100):
            msg = 'Alert - dollar going down - compared to last value -{:.2f}%'.format((1-percent)*100)
            msg = msg+'\nCurrent value: R$ '+str(last_value)
            sub='Dollar Exchange notification'
            Sendmail(sub,msg)
        else:
            # Compare current value to last day value 
            r = list(reversed(InitFile.exchangeoftheday))
            if r:
                percent = last_value/r[0][1]
                if (percent) >= (1+perc_check/100):
                    msg = 'Alert - dollar going up - compared to last day +{:.2f}%'.format((percent-1)*100)
                    msg = msg+'\nCurrent value: R$ '+str(last_value)
                    sub = 'Dollar Exchange notification'
                    Sendmail(sub,msg)
                elif (percent) <= (1-perc_check/100):
                    msg = 'Alert - dollar going down - compared to last day -{:.2f}%'.format((1-percent)*100)
                    msg = msg+'\nCurrent value: R$ '+str(last_value)
                    sub = 'Dollar Exchange notification'
                    Sendmail(sub,msg)
    # time1 = InitFile.ws['B'+str(lastRow)].value
    # timeA = datetime.strptime(time1, "%H:%M:%S")
    # print(timeA)
    now = datetime.now()
    hour = now.hour
    if hour >= 17 and hour < 18:
        r = list(reversed(InitFile.exchangeoftheday))
        percent = last_value/r[0][1]
        
        msg = 'Canadian dollar to Brazilian real: R$ '+str(last_value)
        if percent >1:
            msg = msg + '\n\nUp {:.2f}% compare to yesterday'.format((percent-1)*100) 
        if percent <1:
            msg = msg + '\n\nDown {:.2f}% compare to yesterday'.format((1-percent)*100) 
        for s in range(len(r)):
            msg = msg+'\n\n'+r[s][0]+" - R$ "+str(r[s][1]) 
        sub = 'Daily Dollar Exchange Rate'
        Sendmail(sub,msg)
    # timeB = current_time = now.strftime("%H:%M:%S")
    # print(timeA)
    # print('2 '+ timeB)
    # print(timeA-timeB)

def Sendmail(subject,email_text):
    smtp_server = "smtp.gmail.com"
    port = 587  # For starttls
    email = "froemming@gmail.com"
    password = "pnxnuclzljykzmln"

    # Create a secure SSL context
    context = ssl.create_default_context()

    # Try to log in to server and send email
    BODY = "\r\n".join((
            "From: %s" % email,
            "To: %s" % email,
            "Subject: %s" % subject ,
            "",
            email_text
            ))
    try:
        server = smtplib.SMTP(smtp_server,port)
        server.ehlo() # Can be omitted
        server.starttls(context=context) # Secure the connection
        server.ehlo() # Can be omitted
        server.login(email, password)
        server.sendmail(email, email, BODY)

    except Exception as e:
        # Print any error messages to stdout
        print(e)
    finally:
        server.quit() 
   
def WriteData(date,time,value,api):
    try:
        # # file location 
        # dest_filename = 'data.xlsx'
        # wb = Workbook()

        # # check if file exist, if not, create new file 
        # if not os.path.isfile(dest_filename):
        #      ws = wb.active
        #      ws.title = "Data"
        #      wb.save(filename = dest_filename)
        #      ws.append({'A':'Date','B':'Time','C':'Value'})

        # # load excel shett
        #wb = load_workbook(filename = dest_filename)    
        
        # active sheet
        #ws = wb['Data']

        # add values to datasheet
        InitFile.ws.append({'A':date,'B':time,'C':value,'D':api})
        InitFile.wb.save(filename = InitFile.dest_filename)
        return date,time,value

    except:
        msg="Error writing data to file."
        print(msg)
        return 'error', msg
        
def main(argv): 
    #print ("arguments: ", argv)
    
    # check if argument is passed
    if argv=='-u':
        UpdateCmd()
        CheckChange()
        #Sendmail()
        sys.exit()
    file = InitFile()
    
    root = Tk()
    gui = MainWindow(root)
    gui.ReadLastValue()
    root.mainloop()
    
    
if __name__ == "__main__": 
    if (len(sys.argv)<2):
        argument = ''
    else:
        argument = str(sys.argv[1])
    main(argument)